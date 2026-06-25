#!/usr/bin/env python3
"""
TUDO EM UM COMANDO — baixa, filtra e preenche a planilha de captação.

    python rodar_tudo.py "Clientes_PF_TPC_CPF.xlsx"
    python rodar_tudo.py "Clientes_PF_TPC_CPF.xlsx" --competencia 2025-05 --aba ADVWin

Diferença para o fluxo de 3 passos: este script é "alvo-dirigido". Em vez de
carregar a base inteira (~17 GB) num banco, ele baixa cada arquivo, FILTRA
apenas os clientes da sua planilha e apaga o arquivo em seguida. Uso de disco
fica em poucos GB por vez e a RAM é mínima — roda em um notebook comum.

Requisitos: Python 3.9+, `pip install -r requirements.txt`, internet liberada
para arquivos.receitafederal.gov.br e ~10 GB livres de pico de disco.

Saída: captacao-cnpj/saida/<nome>_PREENCHIDA.xlsx
"""
import sys
import time
import argparse
import zipfile
from pathlib import Path

import requests
import duckdb
import pandas as pd
import openpyxl

from comum import (
    URL_BASE, SHARE_TOKEN, GRUPOS_ARQUIVOS,
    COLUNAS_SOCIOS, COLUNAS_EMPRESAS, COLUNAS_ESTABELECIMENTOS, COLUNAS_DOMINIO,
    sql_normaliza, cpf_mascarado, PORTE, SITUACAO_CADASTRAL,
)

PASTA = Path(__file__).parent
TMP = PASTA / "dados" / "_tmp"
SAIDA = PASTA / "saida"
HEADERS = {"User-Agent": "Mozilla/5.0 (captacao-cnpj)"}
# WebDAV do share público da RFB usa Basic auth: usuário = token, senha vazia.
AUTH = (SHARE_TOKEN, "")

# Alguns CSVs da Receita (visto em Estabelecimentos) trazem bytes de controle C1
# soltos (0x80–0x9F) em campos de contato. O leitor latin-1 do DuckDB rejeita
# esses bytes ("File is not latin-1 encoded"). Trocamos por espaço na extração.
_LIMPA_C1 = bytes(0x20 if 0x80 <= i <= 0x9f else i for i in range(256))


def struct(colunas):
    return "{" + ", ".join(f"'{c}': 'VARCHAR'" for c in colunas) + "}"


def read_csv(glob, colunas):
    return (f"read_csv('{glob}', delim=';', header=false, quote='\"', "
            f"encoding='latin-1', columns={struct(colunas)}, "
            f"ignore_errors=true, null_padding=true)")


def competencia_recente():
    import re
    from datetime import date
    try:
        # WebDAV: PROPFIND com Depth 1 lista as subpastas AAAA-MM da pasta CNPJ.
        r = requests.request("PROPFIND", URL_BASE + "/", auth=AUTH,
                             headers={**HEADERS, "Depth": "1"}, timeout=120)
        meses = sorted(set(re.findall(r"/(\d{4}-\d{2})/", r.text)))
        if meses:
            return meses[-1]
    except Exception:
        pass
    h = date.today()
    return f"{h.year:04d}-{h.month:02d}"


def baixa_e_extrai(base, nome, tentativas=6):
    """Baixa um zip e extrai o CSV (limpando bytes de controle C1); apaga o zip.

    Robusto para downloads longos: retoma de onde parou via HTTP Range quando a
    conexão cai e tenta novamente com espera crescente. Um zip parcial deixado
    por uma execução anterior é reaproveitado (continua o download), não baixado
    de novo do zero.
    """
    TMP.mkdir(parents=True, exist_ok=True)
    zip_path = TMP / nome
    url = f"{base}/{nome}"
    print(f"   baixando {nome} ...", end="", flush=True)
    for tent in range(1, tentativas + 1):
        try:
            ja = zip_path.stat().st_size if zip_path.exists() else 0
            headers = dict(HEADERS)
            modo = "wb"
            if ja:
                headers["Range"] = f"bytes={ja}-"
            with requests.get(url, headers=headers, auth=AUTH,
                              stream=True, timeout=900) as r:
                if ja and r.status_code == 206:
                    modo = "ab"                 # servidor aceitou retomar
                elif r.status_code == 416:
                    pass                         # já temos o arquivo inteiro
                else:
                    modo, ja = "wb", 0          # recomeça do zero
                    r.raise_for_status()
                if r.status_code != 416:
                    with open(zip_path, modo) as f:
                        for chunk in r.iter_content(1 << 20):
                            f.write(chunk)
            with zipfile.ZipFile(zip_path) as z:    # valida o ZIP baixado
                z.namelist()
            break
        except Exception as e:
            espera = min(30, 2 ** tent)
            print(f" [falha {tent}/{tentativas}: {type(e).__name__}; "
                  f"retomando em {espera}s]", end="", flush=True)
            if isinstance(e, zipfile.BadZipFile) and zip_path.exists():
                zip_path.unlink()               # corrompido: recomeça limpo
            time.sleep(espera)
    else:
        raise RuntimeError(f"download falhou após {tentativas} tentativas: {nome}")
    print(f" {zip_path.stat().st_size/1e6:.0f} MB; extraindo...", end="", flush=True)
    with zipfile.ZipFile(zip_path) as z:
        membro = z.namelist()[0]
        destino = TMP / membro
        # extrai já limpando bytes de controle C1 (num único passe de escrita)
        with z.open(membro) as src, open(destino, "wb") as out:
            while True:
                pedaco = src.read(1 << 22)
                if not pedaco:
                    break
                out.write(pedaco.translate(_LIMPA_C1))
    zip_path.unlink()
    print(" ok")
    return destino


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("planilha")
    ap.add_argument("--competencia", default=None)
    ap.add_argument("--aba", default=None)
    ap.add_argument("--col-cliente", default=None)
    ap.add_argument("--col-cpf", default=None)
    args = ap.parse_args()

    competencia = args.competencia or competencia_recente()
    base = f"{URL_BASE}/{competencia}"
    print(f"Competência: {competencia}\n")
    SAIDA.mkdir(exist_ok=True)

    # ---- lê a planilha de clientes ----
    caminho = Path(args.planilha)
    xls = pd.ExcelFile(caminho)
    aba = args.aba or xls.sheet_names[-1]   # última aba (pula "Instruções")
    df = pd.read_excel(caminho, sheet_name=aba, dtype=str)
    cols = {c.lower().strip(): c for c in df.columns}
    col_cliente = args.col_cliente or next((cols[k] for k in ["cliente", "nome"] if k in cols), None)
    col_cpf = args.col_cpf or next((cols[k] for k in ["cpf", "documento"] if k in cols), None)
    if not col_cliente:
        sys.exit(f"Não achei a coluna de cliente em {list(df.columns)}")

    clientes = pd.DataFrame({"cliente": df[col_cliente].fillna("").astype(str).str.strip()})
    clientes = clientes[clientes["cliente"] != ""].drop_duplicates().reset_index(drop=True)
    clientes["cpf_mask"] = (df[col_cpf].map(cpf_mascarado).reindex(clientes.index).values
                            if col_cpf else None)
    print(f"Aba '{aba}': {len(clientes)} clientes ({'com' if col_cpf else 'sem'} CPF).\n")

    # Banco de trabalho EM ARQUIVO: permite retomar de onde parou se a sessão
    # cair no meio (queda de rede, container reiniciado). Cada zip processado é
    # registrado em "_feitos"; numa nova execução os já feitos são pulados.
    DB = PASTA / "dados" / "trabalho.duckdb"
    DB.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(DB))
    con.execute("CREATE TABLE IF NOT EXISTS _feitos (nome VARCHAR PRIMARY KEY);")

    def ja_feito(n):
        return con.execute("SELECT 1 FROM _feitos WHERE nome = ?", [n]).fetchone() is not None

    # "alvo" é recriado a cada execução (barato e determinístico).
    con.register("clientes_in", clientes)
    con.execute(f"CREATE OR REPLACE TABLE alvo AS SELECT cliente, cpf_mask, "
                f"{sql_normaliza('cliente')} AS nome_norm FROM clientes_in;")

    def processa(grupo, ddl, insert_sql, colunas):
        """Baixa cada zip do grupo e roda o INSERT, pulando os já feitos.
        O INSERT e a marca em _feitos vão na mesma transação (atômico)."""
        con.execute(ddl)
        for nome in GRUPOS_ARQUIVOS[grupo]:
            if ja_feito(nome):
                print(f"   {nome}: já processado (pulando)")
                continue
            csv = baixa_e_extrai(base, nome)
            glob = str(csv).replace(chr(92), "/")
            con.execute("BEGIN;")
            con.execute(insert_sql.format(src=read_csv(glob, colunas)))
            con.execute("INSERT INTO _feitos VALUES (?)", [nome])
            con.execute("COMMIT;")
            csv.unlink()

    # ---- 1) varre SÓCIOS filtrando só os nossos clientes ----
    print("== Sócios (filtrando seus clientes) ==")
    processa(
        "Socios",
        "CREATE TABLE IF NOT EXISTS socios_match (cnpj_basico VARCHAR, nome_socio VARCHAR, "
        "cpf_cnpj_socio VARCHAR, qualificacao_socio VARCHAR, data_entrada VARCHAR, "
        "cliente VARCHAR, cpf_mask VARCHAR);",
        "INSERT INTO socios_match "
        "SELECT s.cnpj_basico, s.nome_socio, s.cpf_cnpj_socio, s.qualificacao_socio, "
        "       s.data_entrada_sociedade, a.cliente, a.cpf_mask "
        "FROM {src} s "
        f"JOIN alvo a ON a.nome_norm = {sql_normaliza('s.nome_socio')} "
        "WHERE s.identificador_socio = '2';",
        COLUNAS_SOCIOS,
    )
    n = con.execute("SELECT count(*) FROM socios_match").fetchone()[0]
    print(f"   vínculos de sócios: {n}")

    # marca homônimos (mesmo nome, vários CPFs) e descarta CPF que não confere
    con.execute("""
        CREATE OR REPLACE TABLE socios_f AS
        SELECT *, count(DISTINCT cpf_cnpj_socio) OVER (PARTITION BY cliente) AS qtd_cpfs
        FROM socios_match
        WHERE cpf_mask IS NULL OR cpf_cnpj_socio = cpf_mask;
    """)
    cnpjs = [r[0] for r in con.execute("SELECT DISTINCT cnpj_basico FROM socios_f").fetchall()]
    print(f"\nEmpresas distintas a detalhar: {len(cnpjs)}")
    if not cnpjs:
        print("Nenhum vínculo societário encontrado para esta lista.")
    con.execute("CREATE OR REPLACE TABLE alvo_cnpj AS SELECT DISTINCT cnpj_basico FROM socios_f;")

    # ---- 2) EMPRESAS (capital social, razão, natureza, porte) ----
    print("\n== Empresas (capital social) ==")
    processa(
        "Empresas",
        "CREATE TABLE IF NOT EXISTS emp (cnpj_basico VARCHAR, razao_social VARCHAR, "
        "natureza_juridica VARCHAR, capital_social DOUBLE, porte_empresa VARCHAR);",
        "INSERT INTO emp "
        "SELECT e.cnpj_basico, e.razao_social, e.natureza_juridica, "
        "       TRY_CAST(replace(e.capital_social, ',', '.') AS DOUBLE), e.porte_empresa "
        "FROM {src} e "
        "JOIN alvo_cnpj a ON a.cnpj_basico = e.cnpj_basico;",
        COLUNAS_EMPRESAS,
    )

    # ---- 3) ESTABELECIMENTOS (CNAE/atividade, UF, contato) ----
    print("\n== Estabelecimentos (atividade/contato) ==")
    processa(
        "Estabelecimentos",
        "CREATE TABLE IF NOT EXISTS est (cnpj_basico VARCHAR, cnpj_ordem VARCHAR, cnpj_dv VARCHAR, "
        "situacao_cadastral VARCHAR, cnae_principal VARCHAR, uf VARCHAR, "
        "municipio VARCHAR, ddd_1 VARCHAR, telefone_1 VARCHAR, email VARCHAR);",
        "INSERT INTO est "
        "SELECT t.cnpj_basico, t.cnpj_ordem, t.cnpj_dv, t.situacao_cadastral, "
        "       t.cnae_principal, t.uf, t.municipio, t.ddd_1, t.telefone_1, t.email "
        "FROM {src} t "
        "JOIN alvo_cnpj a ON a.cnpj_basico = t.cnpj_basico "
        "WHERE t.matriz_filial = '1';",
        COLUNAS_ESTABELECIMENTOS,
    )

    # ---- 4) tabelas de domínio (pequenas) ----
    print("\n== Tabelas de referência ==")
    for tabela, grupo, in [("cnaes", "Cnaes"), ("naturezas", "Naturezas"),
                           ("qualificacoes", "Qualificacoes"), ("municipios", "Municipios")]:
        csv = baixa_e_extrai(base, GRUPOS_ARQUIVOS[grupo][0])
        con.execute(f"CREATE OR REPLACE TABLE {tabela} AS SELECT codigo, descricao FROM "
                    f"{read_csv(str(csv).replace(chr(92), '/'), COLUNAS_DOMINIO)};")
        csv.unlink()

    # ---- 5) junta tudo ----
    res = con.execute("""
        SELECT s.cliente, s.cpf_mask, s.cpf_cnpj_socio, s.qtd_cpfs,
               s.cnpj_basico, s.qualificacao_socio, s.data_entrada,
               e.razao_social, e.capital_social, e.porte_empresa, e.natureza_juridica,
               nat.descricao AS natureza_desc, q.descricao AS qualificacao_desc,
               est.cnpj_ordem, est.cnpj_dv, est.situacao_cadastral, est.cnae_principal,
               cn.descricao AS cnae_desc, est.uf, mun.descricao AS municipio_desc,
               est.ddd_1, est.telefone_1, est.email
        FROM socios_f s
        JOIN emp e ON e.cnpj_basico = s.cnpj_basico
        LEFT JOIN est ON est.cnpj_basico = s.cnpj_basico
        LEFT JOIN cnaes cn ON cn.codigo = est.cnae_principal
        LEFT JOIN naturezas nat ON nat.codigo = e.natureza_juridica
        LEFT JOIN qualificacoes q ON q.codigo = s.qualificacao_socio
        LEFT JOIN municipios mun ON mun.codigo = est.municipio
        ORDER BY s.cliente, e.capital_social DESC NULLS LAST;
    """).df()
    con.close()

    # ---- 6) formata e escreve (mesma lógica do cruzar_clientes) ----
    escreve_saida(caminho, aba, col_cliente, clientes, res)

    # limpeza (só após sucesso): apaga CSVs temporários e o banco de trabalho
    try:
        for f in TMP.glob("*"):
            f.unlink()
        TMP.rmdir()
    except Exception:
        pass
    try:
        DB.unlink()
        (DB.parent / (DB.name + ".wal")).unlink(missing_ok=True)
    except Exception:
        pass


def escreve_saida(caminho, aba, col_cliente, clientes, res):
    def fmt_data(v):
        s = str(v) if v is not None else ""
        return f"{s[6:8]}/{s[4:6]}/{s[0:4]}" if len(s) == 8 else ""
    def confianca(r):
        tem = pd.notna(r["cpf_mask"]) and str(r["cpf_mask"]).strip()
        if tem and r["cpf_cnpj_socio"] == r["cpf_mask"]:
            return "ALTA — confirmado por CPF"
        if r["qtd_cpfs"] and r["qtd_cpfs"] > 1:
            return f"BAIXA — possível homônimo ({int(r['qtd_cpfs'])} CPFs c/ esse nome)"
        return "MÉDIA — nome único na base"

    if len(res):
        res["Confiança"] = res.apply(confianca, axis=1)
        res["CNPJ"] = res.apply(lambda r: (
            f"{str(r['cnpj_basico']).zfill(8)[:2]}.{str(r['cnpj_basico']).zfill(8)[2:5]}."
            f"{str(r['cnpj_basico']).zfill(8)[5:8]}/{str(r['cnpj_ordem'] or '0001').zfill(4)}"
            f"-{str(r['cnpj_dv'] or '00').zfill(2)}") if pd.notna(r["cnpj_ordem"]) else "", axis=1)
        res["Entrada na Sociedade"] = res["data_entrada"].map(fmt_data)
        res["Porte"] = res["porte_empresa"].map(lambda v: PORTE.get(str(v), ""))
        res["Situação Cadastral"] = res["situacao_cadastral"].map(lambda v: SITUACAO_CADASTRAL.get(str(v), ""))
        res["Telefone"] = res.apply(lambda r: f"({str(r['ddd_1'] or '').strip()}) {str(r['telefone_1']).strip()}"
                                    if pd.notna(r["telefone_1"]) and str(r["telefone_1"]).strip() else "", axis=1)
        res["Área de Atuação (CNAE)"] = res.apply(
            lambda r: f"{r['cnae_principal']} - {r['cnae_desc']}" if pd.notna(r["cnae_desc"]) else "", axis=1)
        res["Município/UF"] = res.apply(lambda r: f"{r['municipio_desc']}/{r['uf']}" if pd.notna(r["uf"]) else "", axis=1)

    mapa = {"cliente": "Cliente", "Confiança": "Confiança", "razao_social": "Empresa (Razão Social)",
            "CNPJ": "CNPJ", "capital_social": "Capital Social (R$)", "Porte": "Porte",
            "natureza_desc": "Natureza Jurídica", "qualificacao_desc": "Qualificação do Sócio",
            "Entrada na Sociedade": "Entrada na Sociedade", "Situação Cadastral": "Situação Cadastral",
            "Área de Atuação (CNAE)": "Área de Atuação (CNAE)", "Município/UF": "Município/UF",
            "Telefone": "Telefone", "email": "E-mail"}
    detalhe = (res[list(mapa)].rename(columns=mapa) if len(res)
               else pd.DataFrame(columns=list(mapa.values())))

    com = set(res["cliente"]) if len(res) else set()
    sem = clientes[~clientes["cliente"].isin(com)][["cliente"]].rename(
        columns={"cliente": "Cliente sem empresa localizada"})

    melhor = {}
    if len(res):
        for _, r in res.sort_values("capital_social", ascending=False).iterrows():
            melhor.setdefault(r["cliente"], (r["razao_social"], r["capital_social"]))

    out = SAIDA / (caminho.stem + "_PREENCHIDA.xlsx")
    wb = openpyxl.load_workbook(caminho)
    ws = wb[aba]
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci = hdr.index(col_cliente.lower()) + 1 if col_cliente.lower() in hdr else 1
    ce = hdr.index("empresa") + 1 if "empresa" in hdr else None
    cc = hdr.index("capital social") + 1 if "capital social" in hdr else None
    if ce and cc:
        for row in range(2, ws.max_row + 1):
            nome = ws.cell(row=row, column=ci).value
            if nome and str(nome).strip() in melhor:
                emp, cap = melhor[str(nome).strip()]
                ws.cell(row=row, column=ce).value = emp
                ws.cell(row=row, column=cc).value = cap
    for nm, dfx in [("Captacao_Detalhe", detalhe), ("Sem_Match", sem)]:
        if nm in wb.sheetnames:
            del wb[nm]
        nova = wb.create_sheet(nm)
        nova.append(list(dfx.columns))
        for _, r in dfx.iterrows():
            nova.append(["" if pd.isna(v) else v for v in r.tolist()])
    wb.save(out)

    print(f"\n=== RESUMO ===")
    print(f"Clientes com empresa localizada : {len(com)}")
    print(f"Clientes sem match              : {len(sem)}")
    print(f"Vínculos (cliente x empresa)    : {len(detalhe)}")
    print(f"\nArquivo gerado: {out}")


if __name__ == "__main__":
    main()
