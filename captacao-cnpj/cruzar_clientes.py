#!/usr/bin/env python3
"""
Passo 3 — Cruza a planilha de clientes (PF) com a base de sócios e preenche
as empresas, capital social e dados de captação.

Uso:
    python cruzar_clientes.py "Clientes_PF_TPC.xlsx"
    python cruzar_clientes.py "Clientes_PF_TPC.xlsx" --aba ADVWin --col-cliente Cliente --col-cpf CPF

Saídas em captacao-cnpj/saida/:
    - <nome>_PREENCHIDA.xlsx : a planilha original com Empresa/Capital Social
      preenchidos pelo melhor match + abas "Captacao_Detalhe" e "Sem_Match".

Como funciona o match:
    nome do cliente (normalizado) == nome do sócio (normalizado), apenas
    sócios pessoa física. Se a planilha tiver coluna de CPF, o CPF mascarado
    da Receita (***NNNNNN**) é usado para CONFIRMAR o match e descartar
    homônimos automaticamente.
"""
import sys
import argparse
from pathlib import Path

import duckdb
import pandas as pd
import openpyxl

from comum import (
    sql_normaliza,
    normaliza_py,
    cpf_mascarado,
    PORTE,
    SITUACAO_CADASTRAL,
    MATRIZ_FILIAL,
)

PASTA = Path(__file__).parent
BANCO = PASTA / "cnpj.duckdb"
SAIDA = PASTA / "saida"


def detecta_coluna(df, candidatos):
    cols = {c.lower().strip(): c for c in df.columns}
    for cand in candidatos:
        if cand in cols:
            return cols[cand]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("planilha")
    ap.add_argument("--aba", default=None, help="nome da aba (default: primeira)")
    ap.add_argument("--col-cliente", default=None)
    ap.add_argument("--col-cpf", default=None)
    args = ap.parse_args()

    if not BANCO.exists():
        sys.exit(f"Banco não encontrado: {BANCO}. Rode construir_base.py antes.")

    SAIDA.mkdir(exist_ok=True)
    caminho = Path(args.planilha)

    # --- lê a planilha de clientes -----------------------------------------
    xls = pd.ExcelFile(caminho)
    aba = args.aba or xls.sheet_names[0]
    df = pd.read_excel(caminho, sheet_name=aba, dtype=str)

    col_cliente = args.col_cliente or detecta_coluna(df, ["cliente", "nome", "nome do cliente"])
    if not col_cliente:
        sys.exit(f"Não achei a coluna de cliente. Colunas: {list(df.columns)}")
    col_cpf = args.col_cpf or detecta_coluna(df, ["cpf", "cpf/cnpj", "documento"])

    clientes = pd.DataFrame({
        "idx": range(len(df)),
        "cliente": df[col_cliente].fillna("").astype(str).str.strip(),
    })
    clientes = clientes[clientes["cliente"] != ""].copy()
    if col_cpf:
        clientes["cpf_mask"] = df.loc[clientes["idx"], col_cpf].map(cpf_mascarado).values
    else:
        clientes["cpf_mask"] = None

    print(f"Aba '{aba}': {len(clientes)} clientes "
          f"({'com' if col_cpf else 'SEM'} coluna de CPF).")

    # --- cruzamento no DuckDB ----------------------------------------------
    con = duckdb.connect(str(BANCO), read_only=True)
    con.register("clientes_in", clientes)
    con.execute(f"""
        CREATE TEMP TABLE clientes AS
        SELECT idx, cliente, cpf_mask, {sql_normaliza('cliente')} AS nome_norm
        FROM clientes_in;
    """)

    resultado = con.execute("""
        WITH soc AS (
            SELECT s.*,
                   count(DISTINCT s.cpf_cnpj_socio) OVER (PARTITION BY s.nome_norm) AS qtd_cpfs_no_nome
            FROM socios s
            WHERE s.identificador_socio = '2'      -- somente pessoa física
        )
        SELECT
            c.idx,
            c.cliente,
            c.cpf_mask,
            soc.cpf_cnpj_socio,
            soc.qtd_cpfs_no_nome,
            soc.qualificacao_socio,
            soc.data_entrada_sociedade,
            e.cnpj_basico,
            e.razao_social,
            e.capital_social,
            e.porte_empresa,
            e.natureza_juridica,
            nat.descricao        AS natureza_desc,
            q.descricao          AS qualificacao_desc,
            est.cnpj_ordem, est.cnpj_dv, est.matriz_filial,
            est.situacao_cadastral,
            est.cnae_principal,
            cn.descricao         AS cnae_desc,
            est.uf,
            mun.descricao        AS municipio_desc,
            est.ddd_1, est.telefone_1, est.email
        FROM clientes c
        JOIN soc                ON soc.nome_norm = c.nome_norm
        JOIN empresas e         ON e.cnpj_basico = soc.cnpj_basico
        LEFT JOIN estabelecimentos est
               ON est.cnpj_basico = soc.cnpj_basico AND est.matriz_filial = '1'
        LEFT JOIN cnaes cn      ON cn.codigo = est.cnae_principal
        LEFT JOIN naturezas nat ON nat.codigo = e.natureza_juridica
        LEFT JOIN qualificacoes q ON q.codigo = soc.qualificacao_socio
        LEFT JOIN municipios mun ON mun.codigo = est.municipio
        ORDER BY c.idx, e.capital_social DESC NULLS LAST;
    """).df()
    con.close()

    print(f"Matches brutos: {len(resultado)}")

    # --- formatação dos campos ---------------------------------------------
    def fmt_cnpj(r):
        if pd.isna(r["cnpj_basico"]):
            return ""
        b = str(r["cnpj_basico"]).zfill(8)
        o = str(r["cnpj_ordem"] or "0001").zfill(4)
        d = str(r["cnpj_dv"] or "00").zfill(2)
        return f"{b[:2]}.{b[2:5]}.{b[5:8]}/{o}-{d}"

    def fmt_data(v):
        if not v or pd.isna(v) or len(str(v)) != 8:
            return ""
        s = str(v)
        return f"{s[6:8]}/{s[4:6]}/{s[0:4]}"

    def fmt_tel(r):
        if pd.isna(r["telefone_1"]) or not str(r["telefone_1"]).strip():
            return ""
        ddd = str(r["ddd_1"] or "").strip()
        return f"({ddd}) {str(r['telefone_1']).strip()}" if ddd else str(r["telefone_1"]).strip()

    def confianca(r):
        tem_cpf = pd.notna(r["cpf_mask"]) and str(r["cpf_mask"]).strip() != ""
        if tem_cpf and r["cpf_cnpj_socio"] == r["cpf_mask"]:
            return "ALTA — confirmado por CPF"
        if tem_cpf and r["cpf_cnpj_socio"] != r["cpf_mask"]:
            return "DESCARTAR — CPF não confere"
        if r["qtd_cpfs_no_nome"] and r["qtd_cpfs_no_nome"] > 1:
            return f"BAIXA — possível homônimo ({int(r['qtd_cpfs_no_nome'])} CPFs c/ esse nome)"
        return "MÉDIA — nome único na base"

    if len(resultado):
        resultado["CNPJ"] = resultado.apply(fmt_cnpj, axis=1)
        resultado["Confianca"] = resultado.apply(confianca, axis=1)
        resultado["Entrada Sociedade"] = resultado["data_entrada_sociedade"].map(fmt_data)
        resultado["Telefone"] = resultado.apply(fmt_tel, axis=1)
        resultado["Porte"] = resultado["porte_empresa"].map(lambda v: PORTE.get(str(v), ""))
        resultado["Situacao"] = resultado["situacao_cadastral"].map(lambda v: SITUACAO_CADASTRAL.get(str(v), ""))
        resultado["Area de Atuacao (CNAE)"] = resultado.apply(
            lambda r: f"{r['cnae_principal']} - {r['cnae_desc']}" if pd.notna(r["cnae_desc"]) else "", axis=1)
        resultado["Municipio/UF"] = resultado.apply(
            lambda r: f"{r['municipio_desc']}/{r['uf']}" if pd.notna(r["uf"]) else "", axis=1)
        # remove os matches que o CPF descartou
        resultado = resultado[resultado["Confianca"] != "DESCARTAR — CPF não confere"].copy()

    # --- monta abas de saída ------------------------------------------------
    colunas_detalhe = {
        "cliente": "Cliente",
        "Confianca": "Confiança",
        "razao_social": "Empresa (Razão Social)",
        "CNPJ": "CNPJ",
        "capital_social": "Capital Social (R$)",
        "Porte": "Porte",
        "natureza_desc": "Natureza Jurídica",
        "qualificacao_desc": "Qualificação do Sócio",
        "Entrada Sociedade": "Entrada na Sociedade",
        "Situacao": "Situação Cadastral",
        "Area de Atuacao (CNAE)": "Área de Atuação (CNAE)",
        "Municipio/UF": "Município/UF",
        "Telefone": "Telefone",
        "email": "E-mail",
    }
    detalhe = (resultado[list(colunas_detalhe)].rename(columns=colunas_detalhe)
               if len(resultado) else pd.DataFrame(columns=list(colunas_detalhe.values())))

    clientes_com_match = set(resultado["cliente"]) if len(resultado) else set()
    sem_match = clientes[~clientes["cliente"].isin(clientes_com_match)][["cliente"]]
    sem_match = sem_match.rename(columns={"cliente": "Cliente sem empresa localizada"})

    # melhor match por cliente (maior capital social) p/ preencher a planilha original
    melhor = {}
    if len(resultado):
        for _, r in resultado.sort_values("capital_social", ascending=False).iterrows():
            melhor.setdefault(r["cliente"], (r["razao_social"], r["capital_social"]))

    # --- escreve preservando a planilha original ---------------------------
    saida_path = SAIDA / (caminho.stem + "_PREENCHIDA.xlsx")
    wb = openpyxl.load_workbook(caminho)
    ws = wb[aba]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col_emp = headers.index("empresa") + 1 if "empresa" in headers else None
    col_cap = headers.index("capital social") + 1 if "capital social" in headers else None
    col_cli_idx = headers.index(col_cliente.lower()) + 1 if col_cliente.lower() in headers else 3
    if col_emp and col_cap:
        for row in range(2, ws.max_row + 1):
            nome = ws.cell(row=row, column=col_cli_idx).value
            if nome and str(nome).strip() in melhor:
                emp, cap = melhor[str(nome).strip()]
                ws.cell(row=row, column=col_emp).value = emp
                ws.cell(row=row, column=col_cap).value = cap

    # adiciona abas novas
    for nome_aba, dfx in [("Captacao_Detalhe", detalhe), ("Sem_Match", sem_match)]:
        if nome_aba in wb.sheetnames:
            del wb[nome_aba]
        nova = wb.create_sheet(nome_aba)
        nova.append(list(dfx.columns))
        for _, r in dfx.iterrows():
            nova.append(["" if pd.isna(v) else v for v in r.tolist()])

    wb.save(saida_path)

    print(f"\n=== RESUMO ===")
    print(f"Clientes com empresa localizada : {len(clientes_com_match)}")
    print(f"Clientes sem match              : {len(sem_match)}")
    print(f"Linhas de detalhe (cliente x empresa): {len(detalhe)}")
    print(f"\nArquivo gerado: {saida_path}")


if __name__ == "__main__":
    main()
