#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Captação de clientes — Dívida Ativa da União (PGFN)
====================================================

Pipeline para identificar EMPRESAS (pessoas jurídicas) de Minas Gerais — com
foco em Belo Horizonte / Região Metropolitana e na Zona da Mata mineira — que
possuem dívidas relevantes inscritas em Dívida Ativa da União.

Fonte primária (dados públicos):
  PGFN — Dados Abertos da Dívida Ativa
  https://dadosabertos.pgfn.gov.br/

Enriquecimento cadastral (opcional, dados públicos de CNPJ):
  - BrasilAPI:    https://brasilapi.com.br/api/cnpj/v1/<cnpj>
  - Minha Receita: https://minhareceita.org/<cnpj>

O script tem 3 modos de uso:

  1) processar arquivos CSV já baixados da PGFN:
       python captacao_divida_ativa.py processar --input-dir ./pgfn_csv

  2) baixar (best-effort) + processar um trimestre:
       python captacao_divida_ativa.py baixar --trimestre 2025-1 --input-dir ./pgfn_csv
       python captacao_divida_ativa.py processar --input-dir ./pgfn_csv --enriquecer 200

  3) gerar um dataset de DEMONSTRAÇÃO (sintético, sem rede):
       python captacao_divida_ativa.py demo

Saídas (pasta ./saida por padrão):
  - empresas_divida_ativa.csv
  - empresas_divida_ativa.xlsx   (se openpyxl instalado)
  - dados-captacao.json          (consumido pelo dashboard web)

Schema dos CSV da PGFN (separador ';'):
  cpf_cnpj; tipo_pessoa; tipo_devedor; nome_devedor; uf_devedor;
  unidade_responsavel; numero_inscricao; tipo_situacao_inscricao;
  situacao_inscricao; receita_principal; data_inscricao;
  indicador_ajuizado; valor_consolidado

Aviso legal / LGPD: a Dívida Ativa da União é informação pública (Lei nº
12.846/2011 e disponibilização oficial da PGFN). Este pipeline foca em
PESSOAS JURÍDICAS (CNPJ) para fins de prospecção comercial B2B legítima.
Pessoas físicas (CPF) são descartadas por padrão.
"""

from __future__ import annotations

import argparse
import csv
import glob
import io
import json
import os
import random
import re
import sys
import time
import zipfile
from collections import defaultdict
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Configuração geográfica — classificação por unidade_responsavel da PGFN
# ---------------------------------------------------------------------------
# A coluna unidade_responsavel traz a unidade da PGFN/RFB que administra o
# débito (ex.: "DRF BELO HORIZONTE", "DRF JUIZ DE FORA"), o que funciona como
# um bom indicador geográfico do devedor sem precisar cruzar com a base
# gigante de CNPJ da Receita.

REGIOES = {
    "Belo Horizonte / RMBH": [
        "BELO HORIZONTE", "CONTAGEM", "BETIM", "NOVA LIMA", "SABARA",
        "RIBEIRAO DAS NEVES", "SANTA LUZIA", "VESPASIANO", "IBIRITE",
        "LAGOA SANTA", "PEDRO LEOPOLDO", "DERAT", "DEINF",
    ],
    "Zona da Mata": [
        "JUIZ DE FORA", "MURIAE", "UBA", "CATAGUASES", "LEOPOLDINA",
        "VICOSA", "PONTE NOVA", "MANHUACU", "SAO JOAO NEPOMUCENO",
        "BICAS", "RIO POMBA", "CARANGOLA", "VISCONDE DO RIO BRANCO",
    ],
}

# Cidades-polo (para o modo demo / exibição)
CIDADES_BH = ["Belo Horizonte", "Contagem", "Betim", "Nova Lima", "Sabará", "Santa Luzia"]
CIDADES_ZM = ["Juiz de Fora", "Muriaé", "Ubá", "Cataguases", "Leopoldina", "Viçosa", "Ponte Nova"]

VALOR_RELEVANTE_PADRAO = 100_000.0  # R$ — limite default para "dívida relevante"

# Situações em que o débito já NÃO é cobrável — descartadas por padrão.
# (Mantemos "Em cobrança", "Ativa", "Suspensa", "Parcelada", "Garantida", etc.,
# pois ainda representam oportunidade ou contexto relevante.)
SITUACOES_EXTINTAS = ("EXTINT", "CANCELAD", "QUITAD", "BAIXAD", "REMID",
                      "LIQUIDAD", "ANISTIAD")


def classificar_regiao(unidade: str) -> str | None:
    """Retorna o rótulo da região a partir da unidade_responsavel, ou None."""
    u = _strip_acentos((unidade or "").upper())
    for regiao, chaves in REGIOES.items():
        for chave in chaves:
            if chave in u:
                return regiao
    return None


def _strip_acentos(s: str) -> str:
    repl = (
        ("Á", "A"), ("À", "A"), ("Â", "A"), ("Ã", "A"),
        ("É", "E"), ("Ê", "E"), ("Í", "I"),
        ("Ó", "O"), ("Ô", "O"), ("Õ", "O"),
        ("Ú", "U"), ("Ç", "C"),
    )
    for a, b in repl:
        s = s.replace(a, b)
    return s


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def parse_valor(v: str) -> float:
    """Converte '1.234.567,89' (pt-BR) em float."""
    if v is None:
        return 0.0
    v = str(v).strip()
    if not v:
        return 0.0
    v = v.replace(".", "").replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0


def so_digitos(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def formata_cnpj(c: str) -> str:
    c = so_digitos(c).zfill(14)
    if len(c) != 14:
        return c
    return f"{c[0:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"


def eh_pj(tipo_pessoa: str, cpf_cnpj: str) -> bool:
    tp = _strip_acentos((tipo_pessoa or "").upper())
    if "JURIDICA" in tp or tp == "PJ":
        return True
    # fallback: 14 dígitos = CNPJ
    return len(so_digitos(cpf_cnpj)) == 14


# ---------------------------------------------------------------------------
# Leitura dos CSV da PGFN
# ---------------------------------------------------------------------------
COLUNAS = [
    "cpf_cnpj", "tipo_pessoa", "tipo_devedor", "nome_devedor", "uf_devedor",
    "unidade_responsavel", "numero_inscricao", "tipo_situacao_inscricao",
    "situacao_inscricao", "receita_principal", "data_inscricao",
    "indicador_ajuizado", "valor_consolidado",
]


def _abrir_texto(caminho: str):
    """Abre CSV detectando encoding (PGFN historicamente usa latin-1/utf-8)."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            f = open(caminho, "r", encoding=enc, newline="")
            f.readline()
            f.seek(0)
            return f
        except UnicodeDecodeError:
            continue
    return open(caminho, "r", encoding="latin-1", newline="")


def _iter_reader(f):
    """Itera dicts a partir de um stream de texto CSV da PGFN (separador ';')."""
    sample = f.read(4096)
    f.seek(0)
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(f, delimiter=delim)
    header = next(reader, None)
    if not header:
        return
    cols = [_strip_acentos(h.strip().lower()).replace(" ", "_") for h in header]
    tem_header = "cpf_cnpj" in cols or "nome_devedor" in cols
    if not tem_header:
        # arquivo sem cabeçalho: assume a ordem padrão
        cols = COLUNAS
        yield dict(zip(cols, header))
    for row in reader:
        if not row:
            continue
        yield dict(zip(cols, row))


def iter_linhas_csv(caminho: str):
    """Itera as linhas de um arquivo CSV no disco."""
    with _abrir_texto(caminho) as f:
        yield from _iter_reader(f)


def iter_linhas_zip(caminho: str):
    """Itera as linhas de todos os CSV dentro de um .zip (sem extrair)."""
    with zipfile.ZipFile(caminho) as z:
        for nome in z.namelist():
            if not nome.lower().endswith(".csv"):
                continue
            print(f"    · {os.path.basename(caminho)} → {nome}")
            with z.open(nome) as raw:
                # PGFN historicamente usa latin-1; utf-8 como fallback
                try:
                    txt = io.TextIOWrapper(raw, encoding="latin-1", newline="")
                    yield from _iter_reader(txt)
                except UnicodeDecodeError:
                    with z.open(nome) as raw2:
                        txt = io.TextIOWrapper(raw2, encoding="utf-8", newline="")
                        yield from _iter_reader(txt)


def coletar_csvs(input_dir: str) -> list[str]:
    pats = ["*.csv", "**/*.csv", "*.CSV", "**/*.CSV"]
    achados: list[str] = []
    for p in pats:
        achados.extend(glob.glob(os.path.join(input_dir, p), recursive=True))
    return sorted(set(achados))


def coletar_zips(input_dir: str) -> list[str]:
    achados: list[str] = []
    for p in ("*.zip", "**/*.zip", "*.ZIP", "**/*.ZIP"):
        achados.extend(glob.glob(os.path.join(input_dir, p), recursive=True))
    return sorted(set(achados))


# ---------------------------------------------------------------------------
# Agregação por CNPJ
# ---------------------------------------------------------------------------
def processar(input_dir: str, valor_relevante: float, somente_regioes: bool,
              excluir_extintas: bool):
    """Lê todos os CSV/ZIP, filtra e agrega por CNPJ. Retorna lista de empresas."""
    csvs = coletar_csvs(input_dir)
    zips = coletar_zips(input_dir)
    if not csvs and not zips:
        sys.exit(f"Nenhum CSV ou ZIP encontrado em '{input_dir}'. Baixe os dados "
                 f"da PGFN ou use o modo 'demo'.")
    print(f"[i] {len(csvs)} CSV e {len(zips)} ZIP encontrado(s).")

    # fontes: cada item é (rótulo, gerador de linhas)
    fontes = [(c, iter_linhas_csv(c)) for c in csvs]
    fontes += [(z, iter_linhas_zip(z)) for z in zips]

    agg: dict[str, dict] = {}
    total_linhas = 0
    for caminho, linhas in fontes:
        print(f"[i] lendo {os.path.basename(caminho)} ...")
        for row in linhas:
            total_linhas += 1
            if total_linhas % 1_000_000 == 0:
                print(f"    ... {total_linhas:,} linhas")

            uf = (row.get("uf_devedor") or "").strip().upper()
            if uf and uf != "MG":
                continue

            cpf_cnpj = row.get("cpf_cnpj", "")
            if not eh_pj(row.get("tipo_pessoa", ""), cpf_cnpj):
                continue

            unidade = row.get("unidade_responsavel", "")
            regiao = classificar_regiao(unidade)
            if somente_regioes and regiao is None:
                continue

            situacao = (row.get("situacao_inscricao") or "").strip()
            sit_norm = _strip_acentos(situacao.upper())
            if excluir_extintas and any(x in sit_norm for x in SITUACOES_EXTINTAS):
                continue

            cnpj = so_digitos(cpf_cnpj).zfill(14)
            valor = parse_valor(row.get("valor_consolidado"))
            ajuizado = "SIM" in (row.get("indicador_ajuizado") or "").upper()

            e = agg.get(cnpj)
            if e is None:
                e = {
                    "cnpj": cnpj,
                    "razao_social": (row.get("nome_devedor") or "").strip(),
                    "uf": uf or "MG",
                    "unidade_responsavel": unidade.strip(),
                    "regiao": regiao or "Outras (MG)",
                    "divida_total": 0.0,
                    "qtd_inscricoes": 0,
                    "qtd_ajuizadas": 0,
                    "situacoes": set(),
                    "receitas": set(),
                    "data_inscricao_mais_antiga": None,
                }
                agg[cnpj] = e

            e["divida_total"] += valor
            e["qtd_inscricoes"] += 1
            if ajuizado:
                e["qtd_ajuizadas"] += 1
            if situacao:
                e["situacoes"].add(situacao)
            rp = (row.get("receita_principal") or "").strip()
            if rp:
                e["receitas"].add(rp)
            di = _parse_data(row.get("data_inscricao"))
            if di and (e["data_inscricao_mais_antiga"] is None
                       or di < e["data_inscricao_mais_antiga"]):
                e["data_inscricao_mais_antiga"] = di

    print(f"[i] {total_linhas:,} linhas lidas; {len(agg):,} CNPJ(s) agregados.")

    empresas = []
    for e in agg.values():
        if e["divida_total"] < valor_relevante:
            continue
        e["situacoes"] = sorted(e["situacoes"])
        e["receitas"] = sorted(e["receitas"])
        dia = e.pop("data_inscricao_mais_antiga")
        e["data_inscricao_mais_antiga"] = dia.isoformat() if dia else None
        e["ajuizado"] = e["qtd_ajuizadas"] > 0
        e["score"] = calcular_score(e)
        empresas.append(e)

    empresas.sort(key=lambda x: x["divida_total"], reverse=True)
    print(f"[i] {len(empresas):,} empresa(s) com dívida >= "
          f"R$ {valor_relevante:,.2f}.")
    return empresas


def _parse_data(s: str):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def calcular_score(e: dict) -> int:
    """Score de oportunidade 0-100 (heurística de priorização comercial).

    Pondera: tamanho da dívida (principal fator), se há execução fiscal
    ajuizada (urgência → maior propensão a contratar), nº de inscrições
    (complexidade) e antiguidade (risco de prescrição/medidas)."""
    v = e["divida_total"]
    # faixa de valor → 0..60 pts (log-like por faixas)
    if v >= 10_000_000:
        s = 60
    elif v >= 5_000_000:
        s = 52
    elif v >= 1_000_000:
        s = 44
    elif v >= 500_000:
        s = 34
    elif v >= 250_000:
        s = 24
    elif v >= 100_000:
        s = 16
    else:
        s = 8
    if e["qtd_ajuizadas"] > 0:
        s += 20  # execução fiscal em curso = dor aguda
    s += min(e["qtd_inscricoes"], 10)  # até 10 pts por volume de inscrições
    # antiguidade
    di = e.get("data_inscricao_mais_antiga")
    if di:
        try:
            anos = (date.today() - date.fromisoformat(di)).days / 365.25
            s += min(int(anos * 2), 10)
        except (ValueError, TypeError):
            pass
    return max(0, min(100, s))


# ---------------------------------------------------------------------------
# Enriquecimento cadastral (opcional) via API pública de CNPJ
# ---------------------------------------------------------------------------
def enriquecer(empresas: list[dict], limite: int, fonte: str = "brasilapi"):
    try:
        import requests
    except ImportError:
        print("[!] 'requests' não instalado; pulando enriquecimento. "
              "(pip install requests)")
        return
    alvo = empresas[:limite]
    print(f"[i] enriquecendo {len(alvo)} empresa(s) via {fonte} ...")
    for i, e in enumerate(alvo, 1):
        try:
            dados = _consulta_cnpj(requests, e["cnpj"], fonte)
            if dados:
                e.update(dados)
            print(f"    [{i}/{len(alvo)}] {e['cnpj']} ok")
        except Exception as exc:  # noqa: BLE001
            print(f"    [{i}/{len(alvo)}] {e['cnpj']} falhou: {exc}")
        time.sleep(0.8)  # respeitar rate limit das APIs públicas


def _consulta_cnpj(requests, cnpj: str, fonte: str) -> dict:
    if fonte == "minhareceita":
        url = f"https://minhareceita.org/{cnpj}"
    else:
        url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
    r = requests.get(url, timeout=30)
    if r.status_code != 200:
        return {}
    d = r.json()
    tel = d.get("ddd_telefone_1") or d.get("telefone") or ""
    return {
        "nome_fantasia": d.get("nome_fantasia") or "",
        "municipio": (d.get("municipio") or "").title(),
        "bairro": (d.get("bairro") or "").title(),
        "logradouro": (d.get("logradouro") or "").title(),
        "cep": d.get("cep") or "",
        "telefone": tel,
        "email": d.get("email") or "",
        "cnae_principal": _cnae(d),
        "situacao_cadastral": d.get("descricao_situacao_cadastral")
                              or d.get("situacao_cadastral") or "",
        "porte": d.get("porte") or d.get("descricao_porte") or "",
        "data_abertura": d.get("data_inicio_atividade") or d.get("abertura") or "",
        "capital_social": d.get("capital_social") or "",
    }


def _cnae(d: dict) -> str:
    c = d.get("cnae_fiscal_descricao")
    if c:
        return c
    ap = d.get("atividade_principal")
    if isinstance(ap, list) and ap:
        return ap[0].get("text") or ap[0].get("descricao") or ""
    return ""


# ---------------------------------------------------------------------------
# Download (best-effort) de um trimestre da PGFN
# ---------------------------------------------------------------------------
def _baixar_url(requests, url: str, input_dir: str):
    nome = os.path.basename(url.split("?")[0]) or "download.zip"
    destino = os.path.join(input_dir, nome)
    print(f"[i] baixando {nome} ...")
    with requests.get(url, stream=True, timeout=900) as r:
        r.raise_for_status()
        total = int(r.headers.get("content-length", 0))
        baixado = 0
        with open(destino, "wb") as fh:
            for chunk in r.iter_content(1 << 20):
                fh.write(chunk)
                baixado += len(chunk)
                if total:
                    print(f"\r    {baixado/1e6:,.0f} MB / {total/1e6:,.0f} MB",
                          end="", flush=True)
    print(f"\n    salvo em {destino}")


def baixar(trimestre: str, input_dir: str, tipos: list[str],
           urls: list[str] | None = None):
    try:
        import requests
    except ImportError:
        sys.exit("'requests' é necessário para baixar. pip install requests")

    os.makedirs(input_dir, exist_ok=True)

    # 1) URLs explícitas têm prioridade (caminho mais confiável).
    if urls:
        for u in urls:
            _baixar_url(requests, u, input_dir)
        print("[ok] download concluído. Agora rode: processar --input-dir", input_dir)
        return

    # 2) Descoberta automática a partir do índice da PGFN.
    base = "https://dadosabertos.pgfn.gov.br/"
    print(f"[i] descobrindo arquivos do trimestre {trimestre} em {base} ...")
    try:
        idx = requests.get(base, timeout=60).text
    except Exception as exc:  # noqa: BLE001
        sys.exit(f"Falha ao acessar a PGFN: {exc}\n"
                 f"Baixe manualmente em {base} e use 'baixar --url <zip>' "
                 f"ou 'processar --input-dir'.")

    links = re.findall(r'href=["\']([^"\']+\.zip)["\']', idx, flags=re.I)
    ano, t = trimestre.split("-")
    alvo = [l for l in links if ano in l and (f"trimestre_0{t}" in l.lower()
            or f"_{t}_" in l or f"-{t}-" in l or f"_{t}." in l)]
    if not alvo:
        print("[!] Não consegui inferir os links automaticamente. "
              "Links .zip encontrados no índice:")
        for l in links[:60]:
            print("    ", l if l.startswith("http") else base + l.lstrip("/"))
        sys.exit("Copie a(s) URL(s) do trimestre desejado e rode: "
                 "baixar --url <zip1> --url <zip2>")

    for l in alvo:
        url = l if l.startswith("http") else base + l.lstrip("/")
        _baixar_url(requests, url, input_dir)
    print("[ok] download concluído. Agora rode: processar --input-dir", input_dir)


# ---------------------------------------------------------------------------
# Saídas
# ---------------------------------------------------------------------------
def exportar(empresas: list[dict], saida_dir: str, meta: dict):
    os.makedirs(saida_dir, exist_ok=True)
    # JSON (dashboard)
    json_path = os.path.join(saida_dir, "dados-captacao.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "empresas": empresas}, f,
                  ensure_ascii=False, indent=1)
    print(f"[ok] {json_path}")

    # CSV
    csv_path = os.path.join(saida_dir, "empresas_divida_ativa.csv")
    campos = ["score", "razao_social", "nome_fantasia", "cnpj", "divida_total",
              "qtd_inscricoes", "qtd_ajuizadas", "ajuizado", "regiao",
              "municipio", "unidade_responsavel", "situacoes", "receitas",
              "cnae_principal", "situacao_cadastral", "porte", "telefone",
              "email", "data_inscricao_mais_antiga"]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(campos)
        for e in empresas:
            w.writerow([
                e.get("score", ""), e.get("razao_social", ""),
                e.get("nome_fantasia", ""), formata_cnpj(e["cnpj"]),
                f'{e["divida_total"]:.2f}'.replace(".", ","),
                e.get("qtd_inscricoes", ""), e.get("qtd_ajuizadas", ""),
                "Sim" if e.get("ajuizado") else "Não",
                e.get("regiao", ""), e.get("municipio", ""),
                e.get("unidade_responsavel", ""),
                " | ".join(e.get("situacoes", [])),
                " | ".join(e.get("receitas", [])),
                e.get("cnae_principal", ""), e.get("situacao_cadastral", ""),
                e.get("porte", ""), e.get("telefone", ""), e.get("email", ""),
                e.get("data_inscricao_mais_antiga", ""),
            ])
    print(f"[ok] {csv_path}")

    # XLSX (opcional)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Captação"
        ws.append(campos)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1D4ED8")
        for e in empresas:
            ws.append([
                e.get("score", ""), e.get("razao_social", ""),
                e.get("nome_fantasia", ""), formata_cnpj(e["cnpj"]),
                round(e["divida_total"], 2), e.get("qtd_inscricoes", ""),
                e.get("qtd_ajuizadas", ""), "Sim" if e.get("ajuizado") else "Não",
                e.get("regiao", ""), e.get("municipio", ""),
                e.get("unidade_responsavel", ""),
                " | ".join(e.get("situacoes", [])),
                " | ".join(e.get("receitas", [])),
                e.get("cnae_principal", ""), e.get("situacao_cadastral", ""),
                e.get("porte", ""), e.get("telefone", ""), e.get("email", ""),
                e.get("data_inscricao_mais_antiga", ""),
            ])
        ws.freeze_panes = "A2"
        xlsx_path = os.path.join(saida_dir, "empresas_divida_ativa.xlsx")
        wb.save(xlsx_path)
        print(f"[ok] {xlsx_path}")
    except ImportError:
        print("[i] openpyxl não instalado; XLSX não gerado "
              "(pip install openpyxl).")


# ---------------------------------------------------------------------------
# Modo DEMO — dados sintéticos (não usa rede)
# ---------------------------------------------------------------------------
NOMES_FANTASIA = [
    "Metalúrgica", "Têxtil", "Construtora", "Laticínios", "Transportes",
    "Comércio de Alimentos", "Indústria Química", "Agropecuária",
    "Supermercados", "Frigorífico", "Cerâmica", "Móveis", "Auto Peças",
    "Distribuidora de Bebidas", "Confecções", "Calçados", "Mineração",
    "Logística", "Engenharia", "Hospital",
]
SUFIXOS = ["Ltda", "S.A.", "Eireli", "ME", "Comércio e Indústria Ltda"]
SITUACOES = ["Em cobrança", "Ativa em cobrança", "Ajuizada", "Parcelada",
             "Suspensa por decisão judicial"]
RECEITAS = ["IRPJ", "CSLL", "COFINS", "PIS/PASEP", "IPI", "Multas Diversas",
            "Contribuição Previdenciária", "IRRF"]
CNAES = [
    "Fabricação de produtos de metal", "Comércio varejista de mercadorias",
    "Transporte rodoviário de carga", "Fabricação de laticínios",
    "Construção de edifícios", "Fabricação de produtos têxteis",
    "Restaurantes e similares", "Atividades de atenção hospitalar",
]


def gerar_demo(n: int = 120) -> list[dict]:
    random.seed(42)
    empresas = []
    for _ in range(n):
        em_bh = random.random() < 0.6
        regiao = "Belo Horizonte / RMBH" if em_bh else "Zona da Mata"
        cidade = random.choice(CIDADES_BH if em_bh else CIDADES_ZM)
        unidade = ("DRF BELO HORIZONTE" if em_bh else "DRF JUIZ DE FORA")
        nome = (f"{random.choice(NOMES_FANTASIA)} {random.choice(['São','Santa','Boa','Nova','Real','Minas'])} "
                f"{random.choice(['Vista','Esperança','Geral','Forte','do Vale','Center'])} {random.choice(SUFIXOS)}")
        divida = round(random.choice([
            random.uniform(100_000, 500_000),
            random.uniform(500_000, 2_000_000),
            random.uniform(2_000_000, 15_000_000),
        ]), 2)
        qtd = random.randint(1, 22)
        ajuiz = random.randint(0, qtd) if random.random() < 0.6 else 0
        cnpj = "".join(random.choice("0123456789") for _ in range(14))
        di = date(random.randint(2009, 2023), random.randint(1, 12),
                  random.randint(1, 28))
        e = {
            "cnpj": cnpj,
            "razao_social": nome.upper(),
            "nome_fantasia": nome.split(" Ltda")[0].split(" S.A.")[0][:30],
            "uf": "MG",
            "unidade_responsavel": unidade,
            "regiao": regiao,
            "municipio": cidade,
            "divida_total": divida,
            "qtd_inscricoes": qtd,
            "qtd_ajuizadas": ajuiz,
            "ajuizado": ajuiz > 0,
            "situacoes": sorted(set(random.sample(SITUACOES, random.randint(1, 3)))),
            "receitas": sorted(set(random.sample(RECEITAS, random.randint(1, 4)))),
            "cnae_principal": random.choice(CNAES),
            "situacao_cadastral": random.choice(["ATIVA", "ATIVA", "ATIVA", "SUSPENSA"]),
            "porte": random.choice(["ME", "EPP", "DEMAIS", "DEMAIS"]),
            "telefone": f"(31) {random.randint(2000,99999)}-{random.randint(1000,9999)}",
            "email": "contato@" + re.sub(r"[^a-z]", "", nome.lower())[:12] + ".com.br",
            "data_inscricao_mais_antiga": di.isoformat(),
        }
        e["score"] = calcular_score(e)
        empresas.append(e)
    empresas.sort(key=lambda x: x["divida_total"], reverse=True)
    return empresas


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _meta(origem: str, total: int, valor_relevante: float, demo: bool) -> dict:
    return {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "origem": origem,
        "fonte": "PGFN — Dados Abertos da Dívida Ativa da União (https://dadosabertos.pgfn.gov.br/)",
        "total_empresas": total,
        "valor_relevante_min": valor_relevante,
        "foco": "Minas Gerais — Belo Horizonte/RMBH e Zona da Mata",
        "demo": demo,
        "aviso": ("Dados públicos. Foco em pessoas jurídicas (CNPJ) para "
                  "prospecção B2B. Confira sempre a situação atual da inscrição "
                  "no e-CAC/Regularize antes de qualquer abordagem."),
    }


def main():
    p = argparse.ArgumentParser(
        description="Captação de clientes via Dívida Ativa da União (PGFN) — MG/BH/Zona da Mata.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pp = sub.add_parser("processar", help="processa CSVs já baixados da PGFN")
    pp.add_argument("--input-dir", required=True)
    pp.add_argument("--saida", default="saida")
    pp.add_argument("--valor-relevante", type=float, default=VALOR_RELEVANTE_PADRAO)
    pp.add_argument("--todas-regioes", action="store_true",
                    help="não restringir a BH/Zona da Mata (mantém todo MG)")
    pp.add_argument("--incluir-extintas", action="store_true",
                    help="incluir inscrições extintas/quitadas/canceladas (por padrão são descartadas)")
    pp.add_argument("--enriquecer", type=int, default=0, metavar="N",
                    help="enriquecer as N maiores empresas via API de CNPJ")
    pp.add_argument("--fonte-cnpj", choices=["brasilapi", "minhareceita"],
                    default="brasilapi")

    pb = sub.add_parser("baixar", help="baixa um trimestre da PGFN (ou URLs diretas)")
    pb.add_argument("--trimestre", default="", help="ex.: 2025-1 (descoberta automática)")
    pb.add_argument("--input-dir", default="pgfn_csv")
    pb.add_argument("--url", action="append", default=[], dest="urls",
                    help="URL .zip direta da PGFN (pode repetir; ignora --trimestre)")
    pb.add_argument("--tipos", nargs="+",
                    default=["nao_previdenciario", "previdenciario", "fgts"])

    pd = sub.add_parser("demo", help="gera dataset sintético (sem rede)")
    pd.add_argument("--saida", default="saida")
    pd.add_argument("-n", type=int, default=120)

    args = p.parse_args()

    if args.cmd == "baixar":
        if not args.urls and not args.trimestre:
            sys.exit("Informe --trimestre AAAA-T ou ao menos um --url <zip>.")
        baixar(args.trimestre, args.input_dir, args.tipos, args.urls)
        return

    if args.cmd == "demo":
        empresas = gerar_demo(args.n)
        meta = _meta("DEMO (dados sintéticos)", len(empresas),
                     VALOR_RELEVANTE_PADRAO, demo=True)
        exportar(empresas, args.saida, meta)
        print(f"\n[ok] DEMO: {len(empresas)} empresas geradas em '{args.saida}'.")
        return

    if args.cmd == "processar":
        empresas = processar(args.input_dir, args.valor_relevante,
                             somente_regioes=not args.todas_regioes,
                             excluir_extintas=not args.incluir_extintas)
        if args.enriquecer > 0:
            enriquecer(empresas, args.enriquecer, args.fonte_cnpj)
        meta = _meta(f"PGFN CSV ({args.input_dir})", len(empresas),
                     args.valor_relevante, demo=False)
        exportar(empresas, args.saida, meta)


if __name__ == "__main__":
    main()
