#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exportação das planilhas e relatórios (requisitos 12, 15, 25, 26).

Arquivos gerados
----------------
* ``empresas_risco_ambiental_mg.xlsx``  — base principal, priorizada
* ``socios_empresas_risco_ambiental_mg.xlsx`` — quadro societário (separado)
* ``empresas_risco_ambiental_mg.csv``   — base completa, para auditoria
* ``mapa_municipios_mg.json``           — agregados por município (requisito 26)
* ``relatorio_execucao.json`` / ``.md`` — resumo, qualidade e fontes

A planilha de empresas já vem com as colunas de acompanhamento da pesquisa
jurídica manual (``STATUS_PESQUISA_CRIMINAL`` etc.) **em branco**, para
preenchimento pela equipe. O sistema não pesquisa e não preenche nada disso
(requisito 13).
"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime

from comum import formata_cnpj

# ---------------------------------------------------------------------------
# Colunas da planilha principal (requisito 12)
# ---------------------------------------------------------------------------
CAMPOS_EMPRESAS = [
    "ID_EMPRESA", "CNPJ", "CNPJ_BASICO", "RAZAO_SOCIAL", "NOME_FANTASIA",
    "SITUACAO_CADASTRAL", "DATA_SITUACAO_CADASTRAL", "MOTIVO_SITUACAO_CADASTRAL",
    "DATA_ABERTURA", "ANOS_ATIVIDADE", "NATUREZA_JURIDICA",
    "PORTE_RFB", "PORTE_CLASSIFICADO", "MOTIVO_EXCLUSAO_PORTE",
    "OPTANTE_SIMPLES", "OPTANTE_MEI", "CAPITAL_SOCIAL",
    "QTD_ESTABELECIMENTOS_RAIZ", "MATRIZ_FILIAL",
    "CNAE_PRINCIPAL", "DESCRICAO_CNAE_PRINCIPAL", "CNAES_SECUNDARIOS",
    "CNAES_RISCO", "QTD_CNAES_RISCO", "CNAE_MAIOR_EXPOSICAO",
    "DESCRICAO_CNAE_MAIOR_EXPOSICAO", "SETOR_RISCO", "SETORES_ADICIONAIS",
    "CATEGORIA_ANEXO_VIII", "PP_GU_ANEXO_VIII", "NIVEL_RISCO_AMBIENTAL",
    "SCORE_PRIORIDADE", "GRAU_PRIORIDADE", "CLASSIFICACAO_PRIORIDADE",
    "LOGICA_DA_CLASSIFICACAO", "JUSTIFICATIVA_AMBIENTAL",
    "FUNDAMENTO_NORMATIVO",
    "MUNICIPIO", "CODIGO_IBGE", "UF", "ENDERECO", "BAIRRO", "CEP",
    "TELEFONE", "EMAIL",
    "DATA_COLETA", "FONTE", "COMPETENCIA_BASE",
    "STATUS_PESQUISA_CRIMINAL", "OBSERVACOES_PESQUISA_CRIMINAL",
    "STATUS_PESQUISA_MANUAL", "OBSERVACOES",
]

# Tabela PROCESSOS (requisito 15) — criada vazia, como modelo de preenchimento.
CAMPOS_PROCESSOS = [
    "ID_EMPRESA", "CNPJ", "RAZAO_SOCIAL", "NUMERO_PROCESSO_CNJ", "TRIBUNAL",
    "GRAU", "CLASSE", "ASSUNTO", "DATA_DISTRIBUICAO", "POLO", "STATUS",
    "FONTE", "DATA_COLETA", "OBSERVACOES",
]

STATUS_PESQUISA = ["NÃO PESQUISADO", "EM PESQUISA", "SEM RESULTADOS IDENTIFICADOS",
                   "RESULTADO IDENTIFICADO", "NECESSITA ANÁLISE", "DESCARTADO"]


def linha_empresa(e: dict) -> list:
    return [e.get(c, "") for c in CAMPOS_EMPRESAS]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def exportar_csv(empresas: list[dict], caminho: str, log=print) -> None:
    with open(caminho, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CAMPOS_EMPRESAS)
        for e in empresas:
            w.writerow(linha_empresa(e))
    log(f"[ok] {caminho} ({len(empresas):,} linhas)")


def exportar_csv_socios(socios: list[dict], campos: list[str], caminho: str,
                        log=print) -> None:
    with open(caminho, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(campos)
        for s in socios:
            w.writerow([s.get(c, "") for c in campos])
    log(f"[ok] {caminho} ({len(socios):,} linhas)")


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def _estilo_cabecalho(ws, cor: str):
    from openpyxl.styles import Alignment, Font, PatternFill
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=cor)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def exportar_xlsx(empresas: list[dict], socios: list[dict], campos_socios: list[str],
                  matriz, meta: dict, agregados: list[dict],
                  caminho_empresas: str, caminho_socios: str, log=print) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("[i] openpyxl não instalado; XLSX não gerado (pip install openpyxl).")
        return False

    cores_grau = {"A": "FEE2E2", "B": "FEF3C7", "C": "F1F5F9"}

    # ---------------- planilha de EMPRESAS ------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "EMPRESAS"
    ws.append(CAMPOS_EMPRESAS)
    _estilo_cabecalho(ws, "14532D")

    col_grau = CAMPOS_EMPRESAS.index("GRAU_PRIORIDADE") + 1
    for e in empresas:
        ws.append(linha_empresa(e))
        ws.cell(row=ws.max_row, column=col_grau).fill = PatternFill(
            "solid", fgColor=cores_grau.get(e.get("GRAU_PRIORIDADE", ""), "FFFFFF"))

    larguras = {
        "RAZAO_SOCIAL": 42, "NOME_FANTASIA": 28, "CNPJ": 20,
        "DESCRICAO_CNAE_PRINCIPAL": 45, "DESCRICAO_CNAE_MAIOR_EXPOSICAO": 45,
        "CNAES_SECUNDARIOS": 30, "CNAES_RISCO": 26, "SETOR_RISCO": 32,
        "SETORES_ADICIONAIS": 30, "CATEGORIA_ANEXO_VIII": 34,
        "LOGICA_DA_CLASSIFICACAO": 90, "JUSTIFICATIVA_AMBIENTAL": 80,
        "FUNDAMENTO_NORMATIVO": 60, "ENDERECO": 40, "MUNICIPIO": 22,
        "MOTIVO_EXCLUSAO_PORTE": 40, "FONTE": 34, "EMAIL": 28,
        "OBSERVACOES_PESQUISA_CRIMINAL": 34, "OBSERVACOES": 26,
        "STATUS_PESQUISA_CRIMINAL": 24, "STATUS_PESQUISA_MANUAL": 22,
    }
    for i, campo in enumerate(CAMPOS_EMPRESAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(campo, 15)
    ws.auto_filter.ref = ws.dimensions

    # ---- aba de leitura obrigatória ------------------------------------
    wsa = wb.create_sheet("LEIA-ME")
    avisos = [
        ["COMO LER ESTA BASE"],
        [""],
        ["O que esta planilha É"],
        ["Relação de empresas de Minas Gerais cuja ATIVIDADE ECONÔMICA DECLARADA "
         "está sujeita a licenciamento, controle e fiscalização ambiental, "
         "selecionadas a partir de dados públicos oficiais para posterior "
         "verificação jurídica individualizada."],
        [""],
        ["O que esta planilha NÃO É"],
        ["NÃO é lista de infratores, investigados, denunciados ou condenados."],
        ["NÃO afirma nem sugere que qualquer empresa aqui listada tenha praticado "
         "infração administrativa, ilícito civil ou crime."],
        ["NÃO contém resultado de pesquisa em tribunais, Ministério Público ou "
         "polícias — essa etapa é manual e posterior (colunas próprias, em branco)."],
        [""],
        ["O que o SCORE significa"],
        ["Prioridade de VERIFICAÇÃO, combinando a exposição ambiental potencial da "
         "atividade declarada com a adequação ao perfil comercial-alvo (pequeno e "
         "médio porte). Não é probabilidade de irregularidade."],
        ["A coluna LOGICA_DA_CLASSIFICACAO detalha, empresa a empresa, de onde veio "
         "cada ponto do score."],
        [""],
        ["Cuidados de uso"],
        ["Os dados cadastrais têm a defasagem da competência indicada na coluna "
         "COMPETENCIA_BASE. Confirme a situação atual antes de qualquer abordagem."],
        ["A planilha de sócios é arquivo SEPARADO e contém dado pessoal: trate "
         "conforme a LGPD e restrinja o acesso ao necessário."],
        ["Ao registrar resultado de pesquisa criminal, use exclusivamente fonte "
         "oficial e anote a referência na coluna de observações."],
        [""],
        ["Valores válidos para STATUS_PESQUISA_CRIMINAL"],
        [" | ".join(STATUS_PESQUISA)],
    ]
    for linha in avisos:
        wsa.append(linha)
    wsa.column_dimensions["A"].width = 118
    for row in wsa.iter_rows(min_row=1, max_row=wsa.max_row, max_col=1):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    for r in (1, 3, 6, 11, 16, 21):
        if r <= wsa.max_row:
            wsa.cell(row=r, column=1).font = Font(bold=True, size=12,
                                                  color="14532D")

    # ---- aba MATRIZ DE CNAES -------------------------------------------
    wsm = wb.create_sheet("MATRIZ_CNAE")
    cols_m = ["CNAE", "DESCRICAO", "SETOR", "CATEGORIA_ANEXO_VIII", "PP_GU",
              "EXPOSICAO_AMBIENTAL", "PESO", "JUSTIFICATIVA", "FONTES"]
    wsm.append(cols_m)
    _estilo_cabecalho(wsm, "1E3A8A")
    for cod, item in sorted(matriz.cnaes.items()):
        wsm.append([cod, item["descricao"], item["setor"],
                    item["categoria_anexo_viii"], item["pp_gu_anexo_viii"],
                    item["exposicao_ambiental"], item["peso"],
                    item["justificativa"], " • ".join(item["fontes"])])
    for i, w in enumerate([12, 52, 32, 34, 10, 14, 7, 95, 80], 1):
        wsm.column_dimensions[get_column_letter(i)].width = w
    wsm.auto_filter.ref = wsm.dimensions

    # ---- aba MAPA (agregados por município) ----------------------------
    wsg = wb.create_sheet("MAPA_MUNICIPIOS")
    cols_g = ["MUNICIPIO", "CODIGO_IBGE", "QTD_EMPRESAS", "SCORE_MEDIO",
              "QTD_GRAU_A", "QTD_GRAU_B", "QTD_GRAU_C", "SETORES_PREDOMINANTES"]
    wsg.append(cols_g)
    _estilo_cabecalho(wsg, "0F766E")
    for a in agregados:
        wsg.append([a["municipio"], a["codigo_ibge"], a["qtd_empresas"],
                    a["score_medio"], a["grau_a"], a["grau_b"], a["grau_c"],
                    " | ".join(a["setores_predominantes"])])
    for i, w in enumerate([26, 14, 14, 14, 12, 12, 12, 60], 1):
        wsg.column_dimensions[get_column_letter(i)].width = w
    wsg.auto_filter.ref = wsg.dimensions

    # ---- aba PROCESSOS (modelo vazio, requisito 15) --------------------
    wsp = wb.create_sheet("PROCESSOS")
    wsp.append(CAMPOS_PROCESSOS)
    _estilo_cabecalho(wsp, "7C2D12")
    wsp.append(["", "", "", "", "", "", "", "", "", "", "", "", "",
                "Tabela criada VAZIA. Preencher somente quando um processo for "
                "efetivamente identificado em fonte oficial. Não existe 'número "
                "CNJ' no cadastro da Receita Federal: o identificador da empresa "
                "é o CNPJ; o número CNJ só existe se houver processo."])
    for i, w in enumerate([12, 20, 40, 28, 14, 8, 22, 30, 18, 14, 16, 24, 14, 90], 1):
        wsp.column_dimensions[get_column_letter(i)].width = w

    # ---- aba FONTES -----------------------------------------------------
    wsf = wb.create_sheet("FONTES")
    wsf.append(["FONTE", "DESCRICAO", "CAMPOS EXTRAÍDOS", "DATA DE ACESSO",
                "LIMITAÇÕES"])
    _estilo_cabecalho(wsf, "334155")
    for f in meta.get("fontes", []):
        wsf.append([f["fonte"], f["descricao"], f["campos"], f["data_acesso"],
                    f["limitacoes"]])
    for i, w in enumerate([38, 60, 60, 16, 78], 1):
        wsf.column_dimensions[get_column_letter(i)].width = w

    wb.save(caminho_empresas)
    log(f"[ok] {caminho_empresas} ({len(empresas):,} empresas)")

    # ---------------- planilha de SÓCIOS (arquivo separado) -------------
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "SOCIOS"
    ws2.append(campos_socios)
    _estilo_cabecalho(ws2, "581C87")
    for s in socios:
        ws2.append([s.get(c, "") for c in campos_socios])
    larg2 = {"NOME_SOCIO": 40, "RAZAO_SOCIAL": 40, "QUALIFICACAO": 30,
             "CNPJ": 20, "REPRESENTANTE_LEGAL": 32, "FONTE": 34,
             "QUALIFICACAO_REPRESENTANTE": 28, "MUNICIPIO": 22,
             "OBSERVACOES": 26}
    for i, campo in enumerate(campos_socios, 1):
        ws2.column_dimensions[get_column_letter(i)].width = larg2.get(campo, 16)
    ws2.auto_filter.ref = ws2.dimensions

    ws2a = wb2.create_sheet("LEIA-ME")
    for linha in [
        ["QUADRO SOCIETÁRIO — USO RESTRITO"],
        [""],
        ["Este arquivo contém DADO PESSOAL e é mantido SEPARADO da base "
         "empresarial justamente por isso. Restrinja o acesso ao estritamente "
         "necessário e não o publique."],
        [""],
        ["Origem: Quadro de Sócios e Administradores dos Dados Abertos do CNPJ "
         "(Receita Federal). O registro empresarial é público por força do art. "
         "29 da Lei nº 8.934/1994."],
        ["O CPF já vem MASCARADO pela própria Receita Federal (formato "
         "***XXXXXX**). Este pipeline não desmascara, não completa e não cruza "
         "esse dado com nenhuma outra base."],
        ["Não são coletados endereço residencial, telefone pessoal ou e-mail "
         "pessoal de sócio."],
        [""],
        ["A coluna E_ADMINISTRADOR sinaliza qualificação com poder de "
         "administração. É indicação cadastral de quem dirige a atividade — NÃO "
         "é atribuição de responsabilidade a quem quer que seja."],
        ["Relacione com a base principal por ID_EMPRESA ou CNPJ."],
    ]:
        ws2a.append(linha)
    ws2a.column_dimensions["A"].width = 118
    for row in ws2a.iter_rows(min_row=1, max_row=ws2a.max_row, max_col=1):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    ws2a.cell(row=1, column=1).font = Font(bold=True, size=12, color="581C87")

    wb2.save(caminho_socios)
    log(f"[ok] {caminho_socios} ({len(socios):,} sócios)")
    return True


# ---------------------------------------------------------------------------
# Agregados por município (requisito 26 — preparação do mapa)
# ---------------------------------------------------------------------------
def agregar_municipios(empresas: list[dict]) -> list[dict]:
    agg: dict[tuple[str, str], dict] = {}
    for e in empresas:
        chave = (e.get("MUNICIPIO", ""), e.get("CODIGO_IBGE", ""))
        if not chave[0]:
            continue
        a = agg.setdefault(chave, {
            "municipio": chave[0], "codigo_ibge": chave[1],
            "qtd_empresas": 0, "soma_score": 0, "grau_a": 0, "grau_b": 0,
            "grau_c": 0, "setores": {},
        })
        a["qtd_empresas"] += 1
        a["soma_score"] += e.get("SCORE_PRIORIDADE", 0)
        grau = e.get("GRAU_PRIORIDADE", "C")
        a[f"grau_{grau.lower()}"] = a.get(f"grau_{grau.lower()}", 0) + 1
        setor = e.get("SETOR_RISCO", "")
        if setor:
            a["setores"][setor] = a["setores"].get(setor, 0) + 1

    saida = []
    for a in agg.values():
        top = sorted(a["setores"].items(), key=lambda kv: -kv[1])[:3]
        saida.append({
            "municipio": a["municipio"],
            "codigo_ibge": a["codigo_ibge"],
            "qtd_empresas": a["qtd_empresas"],
            "score_medio": round(a["soma_score"] / a["qtd_empresas"], 1),
            "grau_a": a["grau_a"], "grau_b": a["grau_b"], "grau_c": a["grau_c"],
            "setores_predominantes": [f"{s} ({n})" for s, n in top],
        })
    saida.sort(key=lambda x: -x["qtd_empresas"])
    return saida


def exportar_mapa(agregados: list[dict], meta: dict, caminho: str,
                  log=print) -> None:
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump({
            "meta": {
                "gerado_em": meta["gerado_em"],
                "uf": "MG",
                "descricao": ("Agregados por município para geração de mapa "
                              "temático. O código IBGE permite o cruzamento "
                              "direto com as malhas territoriais do IBGE."),
                "fonte_geometria_sugerida": (
                    "IBGE — Malhas Territoriais / API de Localidades "
                    "(https://servicodados.ibge.gov.br/api/v1/localidades)"),
                "total_municipios": len(agregados),
            },
            "municipios": agregados,
        }, f, ensure_ascii=False, indent=1)
    log(f"[ok] {caminho} ({len(agregados)} municípios)")


# ---------------------------------------------------------------------------
# Relatório final (requisito 25)
# ---------------------------------------------------------------------------
def exportar_relatorio(rel: dict, caminho_json: str, caminho_md: str,
                       log=print) -> None:
    with open(caminho_json, "w", encoding="utf-8") as f:
        json.dump(rel, f, ensure_ascii=False, indent=1)
    log(f"[ok] {caminho_json}")

    e = rel["estatisticas"]
    L = [
        "# Relatório de execução — Empresas de MG com exposição ambiental potencial",
        "",
        f"**Gerado em:** {rel['gerado_em']}  ",
        f"**Competência da base da Receita Federal:** {rel['competencia']}  ",
        f"**Matriz de CNAEs:** {rel['matriz_cnaes']['total']} subclasses "
        f"(de {rel['matriz_cnaes']['total_oficiais']} oficiais)",
        "",
        "> Esta base identifica **exposição potencial a risco ambiental em razão "
        "da atividade econômica declarada**. Não afirma, não sugere e não "
        "presume a prática de infração ou de crime por nenhuma das empresas "
        "listadas.",
        "",
        "## 1. Funil de seleção",
        "",
        "| Etapa | Registros |",
        "|---|---:|",
        f"| Estabelecimentos lidos (Brasil) | {e['estabelecimentos_lidos']:,} |",
        f"| Estabelecimentos em Minas Gerais | {e['estabelecimentos_mg']:,} |",
        f"| Com CNAE de risco (principal ou secundário) | {e['com_cnae_risco']:,} |",
        f"| Excluídos por situação cadastral | {e['excluidos_situacao']:,} |",
        f"| Excluídos por porte | {e['excluidos_porte']:,} |",
        f"| **Base final** | **{e['base_final']:,}** |",
        "",
        f"Destes, **{e['porte_nao_identificado']:,}** ficaram com porte não "
        f"identificado (mantidos na base e sinalizados, conforme o requisito).",
        "",
        "## 2. Distribuição por grau de prioridade",
        "",
        "| Grau | Empresas |",
        "|---|---:|",
    ]
    for grau, n in sorted(e["por_grau"].items()):
        L.append(f"| {grau} | {n:,} |")

    L += ["", "## 3. Distribuição por setor", "", "| Setor | Empresas |", "|---|---:|"]
    for setor, n in sorted(e["por_setor"].items(), key=lambda kv: -kv[1]):
        L.append(f"| {setor} | {n:,} |")

    L += ["", "## 4. Municípios com maior concentração", "",
          "| Município | Empresas | Score médio | Grau A |", "|---|---:|---:|---:|"]
    for m in rel["top_municipios"][:25]:
        L.append(f"| {m['municipio']} | {m['qtd_empresas']:,} | "
                 f"{m['score_medio']} | {m['grau_a']:,} |")

    L += ["", "## 5. Qualidade da base", "", "```", rel["qualidade_resumo"], "```"]

    L += ["", "## 6. Fontes utilizadas", ""]
    for f in rel["fontes"]:
        L += [f"### {f['fonte']}",
              f"- **Descrição:** {f['descricao']}",
              f"- **Campos extraídos:** {f['campos']}",
              f"- **Data de acesso:** {f['data_acesso']}",
              f"- **Limitações:** {f['limitacoes']}", ""]

    L += ["## 7. Próxima etapa (manual)", "",
          "A base está preparada para a pesquisa jurídica individualizada. As "
          "colunas `STATUS_PESQUISA_CRIMINAL` e `OBSERVACOES_PESQUISA_CRIMINAL` "
          "estão em branco para preenchimento pela equipe.", "",
          "A ausência de resultado em uma fonte **não significa** inexistência "
          "de procedimento: significa apenas que aquela fonte, naquela consulta, "
          "nada retornou.", ""]

    with open(caminho_md, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    log(f"[ok] {caminho_md}")
