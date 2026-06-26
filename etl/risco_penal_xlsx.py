#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera o relatório XLSX de risco penal-tributário com o QSA (sócios).

Combina:
  - dados-captacao-com-risco-penal.json  (campo risco_penal por empresa)
  - qsa_cache.json                       ({cnpj: [socios]} da BrasilAPI)

USO INTERNO — não publicar (dado pessoal de sócios; inferência penal).
"""
from __future__ import annotations
import argparse, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def fcnpj(c: str) -> str:
    c = "".join(ch for ch in c if ch.isdigit()).zfill(14)
    return f"{c[0:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"


def fmt_socios(socios: list[dict], so_admin: bool) -> str:
    linhas = []
    for s in socios:
        if so_admin and not s.get("admin"):
            continue
        nome = s.get("nome", "").strip()
        if not nome:
            continue
        q = s.get("qualificacao", "").strip()
        cpf = s.get("cpf_cnpj", "").strip()
        peca = nome
        if q:
            peca += f" ({q})"
        if cpf:
            peca += f" [{cpf}]"
        linhas.append(peca)
    return "\n".join(linhas)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--risco", default="./dados-captacao-com-risco-penal.json")
    p.add_argument("--qsa", default="./qsa_cache.json")
    p.add_argument("--saida", default="./risco_penal.xlsx")
    args = p.parse_args()

    emp = json.load(open(args.risco, encoding="utf-8"))["empresas"]
    qsa = json.load(open(args.qsa, encoding="utf-8"))

    ordem = {"Alto": 0, "Médio": 1, "Médio-baixo": 2, "Indeterminado": 3}
    emp.sort(key=lambda e: (ordem.get(e["risco_penal"]["nivel"], 9),
                            -e["risco_penal"]["score"],
                            -e.get("divida_total", 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Risco Penal"
    cols = ["Nível", "Score", "Razão social", "CNPJ", "Dívida (R$)", "Região",
            "Município", "Telefone", "Administradores (resp. penal provável)",
            "Todos os sócios (QSA)", "Tipificações possíveis", "Fundamentos",
            "Observações"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7F1D1D")
        c.alignment = Alignment(vertical="top", wrap_text=True)

    cores = {"Alto": "FECACA", "Médio": "FEF9C3", "Médio-baixo": "E0F2FE",
             "Indeterminado": "F3F4F6"}
    sem_qsa = 0
    for e in emp:
        rp = e["risco_penal"]
        socios = qsa.get(e["cnpj"], [])
        if not socios:
            sem_qsa += 1
        ws.append([
            rp["nivel"], rp["score"], e.get("razao_social", ""),
            fcnpj(e["cnpj"]), round(e.get("divida_total", 0), 2),
            e.get("regiao", ""), e.get("municipio", ""), e.get("telefone", ""),
            fmt_socios(socios, so_admin=True) or "—",
            fmt_socios(socios, so_admin=False) or "—",
            " | ".join(rp["tipificacoes"]),
            " | ".join(rp["fundamentos"]),
            " | ".join(rp["observacoes"]),
        ])
        row = ws.max_row
        ws.cell(row=row, column=1).fill = PatternFill(
            "solid", fgColor=cores.get(rp["nivel"], "FFFFFF"))
        for col in (9, 10, 11, 12, 13):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True)

    widths = [13, 7, 36, 20, 16, 18, 15, 15, 40, 48, 50, 48, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(args.saida)
    print(f"[ok] {args.saida}: {len(emp)} linhas; {sem_qsa} sem QSA.")


if __name__ == "__main__":
    main()
